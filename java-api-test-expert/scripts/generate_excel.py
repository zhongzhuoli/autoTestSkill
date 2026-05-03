#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 test_model.json 生成 test_cases.xlsx 工作簿

用法: python generate_excel.py <test_model.json> [输出目录]
依赖: pip install openpyxl
"""

import json
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))
from shared_utils import load_model, get_output_prefix, ensure_output_dir

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("错误: 需要安装 openpyxl。请运行: pip install openpyxl")
    sys.exit(1)


# 样式定义
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CELL_FONT = Font(name="Microsoft YaHei", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data(ws, row_count, col_count):
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER


def _display_width(text):
    """计算文本的显示宽度，CJK 字符算 2，其他算 1"""
    width = 0
    for ch in str(text):
        if '一' <= ch <= '鿿' or '　' <= ch <= '〿' or '＀' <= ch <= '￯':
            width += 2
        else:
            width += 1
    return width


def auto_width(ws, col_count, max_width=50):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, _display_width(val))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, max_width)


# ====== Sheet 1: 扫描摘要 ======
def write_summary_sheet(wb, model):
    ws = wb.create_sheet("扫描摘要")
    data = [
        ["指标", "值"],
        ["扫描模式", model.get("scanMode", "")],
        ["项目名称", model.get("projectName", "")],
        ["模块名称", model.get("moduleName", "")],
        ["Controller 名称", model.get("controllerName", "")],
        ["扫描路径", model.get("scanPath", "")],
        ["baseUrl", model.get("baseUrl", "")],
        ["context-path", model.get("contextPath", "")],
        ["是否识别鉴权", str(model.get("authDetected", False))],
    ]
    stats = model.get("statistics", {})
    data.append(["", ""])
    data.append(["统计指标", "数量"])
    for k, v in stats.items():
        data.append([k, str(v)])

    for r, row in enumerate(data, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    style_header(ws, 2)
    style_data(ws, len(data), 2)
    auto_width(ws, 2)


# ====== Sheet 2: 接口清单 ======
def write_api_list_sheet(wb, model):
    ws = wb.create_sheet("接口清单")
    headers = ["序号", "模块", "Controller", "接口名称", "请求方法", "路径",
               "入参类型", "是否鉴权", "是否危险接口", "解析状态"]
    ws.append(headers)

    for i, api in enumerate(model.get("apis", []), 1):
        ws.append([
            i,
            api.get("moduleName", ""),
            api.get("controllerName", ""),
            api.get("apiName", ""),
            api.get("method", ""),
            api.get("path", ""),
            api.get("requestBodyType", ""),
            str(api.get("authRequired", "")),
            str(api.get("dangerous", "")),
            api.get("parseStatus", ""),
        ])

    style_header(ws, len(headers))
    style_data(ws, len(model.get("apis", [])) + 1, len(headers))
    auto_width(ws, len(headers))


# ====== Sheet 3: 测试用例 ======
TEST_CASE_COLUMNS = [
    "caseId", "moduleName", "controllerName", "apiName", "method", "path",
    "caseName", "caseType", "field", "description", "precondition",
    "requestHeaders", "pathVariables", "queryParams", "requestBody",
    "expectedHttpStatus", "expectedBizCode", "expectedMessageContains",
    "expectedJsonPath", "expectedResult", "priority", "riskLevel",
    "enabled", "needManualConfirm", "manualConfirmReason", "remark"
]


def write_test_cases_sheet(wb, model):
    ws = wb.create_sheet("测试用例")
    ws.append(TEST_CASE_COLUMNS)

    for tc in model.get("testCases", []):
        row = []
        for col in TEST_CASE_COLUMNS:
            val = tc.get(col, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            elif isinstance(val, bool):
                val = str(val)
            row.append(val)
        ws.append(row)

    style_header(ws, len(TEST_CASE_COLUMNS))
    style_data(ws, len(model.get("testCases", [])) + 1, len(TEST_CASE_COLUMNS))
    auto_width(ws, len(TEST_CASE_COLUMNS))


# ====== Sheet 4: JMeter 参数化数据 ======
def write_jmeter_params_sheet(wb, model):
    ws = wb.create_sheet("JMeter 参数化数据")
    csv_fields = [
        "caseId", "caseName", "caseType", "moduleName", "method", "path",
        "headers", "queryParams", "pathVariables", "body",
        "expectedHttpStatus", "expectedBizCode", "expectedMessageContains",
        "enabled", "riskLevel"
    ]
    ws.append(csv_fields)

    for tc in model.get("testCases", []):
        row = []
        headers_val = tc.get("requestHeaders", {})
        if not headers_val:
            headers_val = {"Content-Type": "application/json", "Authorization": "Bearer ${token}"}

        field_map = {
            "headers": json.dumps(headers_val, ensure_ascii=False),
            "queryParams": json.dumps(tc.get("queryParams", {}), ensure_ascii=False),
            "pathVariables": json.dumps(tc.get("pathVariables", {}), ensure_ascii=False),
            "body": json.dumps(tc.get("requestBody"), ensure_ascii=False) if tc.get("requestBody") else "",
            "expectedHttpStatus": str(tc.get("expectedHttpStatus", "")),
            "expectedBizCode": str(tc.get("expectedBizCode", "")),
            "enabled": str(tc.get("enabled", True)).lower(),
        }

        for field in csv_fields:
            if field in field_map:
                row.append(field_map[field])
            else:
                row.append(str(tc.get(field, "")))
        ws.append(row)

    style_header(ws, len(csv_fields))
    style_data(ws, len(model.get("testCases", [])) + 1, len(csv_fields))
    auto_width(ws, len(csv_fields))


# ====== Sheet 5: 风险清单 ======
def write_risks_sheet(wb, model):
    ws = wb.create_sheet("风险清单")
    headers = ["序号", "接口", "风险类型", "原因", "建议"]
    ws.append(headers)

    for i, risk in enumerate(model.get("risks", []), 1):
        ws.append([
            i,
            risk.get("api", ""),
            risk.get("riskType", ""),
            risk.get("reason", ""),
            risk.get("suggestion", ""),
        ])

    style_header(ws, len(headers))
    style_data(ws, len(model.get("risks", [])) + 1, len(headers))
    auto_width(ws, len(headers))


# ====== Sheet 6: 需人工确认项 ======
def write_manual_confirm_sheet(wb, model):
    ws = wb.create_sheet("需人工确认项")
    headers = ["编号", "类型", "位置", "问题", "建议"]
    ws.append(headers)

    for item in model.get("manualConfirmItems", []):
        ws.append([
            item.get("id", ""),
            item.get("type", ""),
            item.get("location", ""),
            item.get("problem", ""),
            item.get("suggestion", ""),
        ])

    style_header(ws, len(headers))
    style_data(ws, len(model.get("manualConfirmItems", [])) + 1, len(headers))
    auto_width(ws, len(headers))


# ====== Sheet 7: 危险接口 ======
def write_dangerous_apis_sheet(wb, model):
    ws = wb.create_sheet("危险接口")
    headers = ["序号", "接口", "方法", "路径", "危险关键词", "风险等级", "压测排除"]
    ws.append(headers)

    for i, api in enumerate(model.get("dangerousApis", []), 1):
        ws.append([
            i,
            api.get("apiName", ""),
            api.get("method", ""),
            api.get("path", ""),
            api.get("dangerKeyword", ""),
            api.get("riskLevel", "high"),
            str(api.get("excludeFromPerf", True)),
        ])

    style_header(ws, len(headers))
    style_data(ws, len(model.get("dangerousApis", [])) + 1, len(headers))
    auto_width(ws, len(headers))


def main():
    if len(sys.argv) < 2:
        print("用法: python generate_excel.py <test_model.json> [输出目录]")
        sys.exit(1)

    model_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(model_path)

    ensure_output_dir(output_dir)

    model = load_model(model_path)
    prefix = get_output_prefix(model)
    xlsx_path = os.path.join(output_dir, f"{prefix}_test_cases.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, model)
    write_api_list_sheet(wb, model)
    write_test_cases_sheet(wb, model)
    write_jmeter_params_sheet(wb, model)
    write_risks_sheet(wb, model)
    write_manual_confirm_sheet(wb, model)
    write_dangerous_apis_sheet(wb, model)

    wb.save(xlsx_path)
    print(f"  生成: {xlsx_path}")
    print("Excel 工作簿生成完成")


if __name__ == "__main__":
    main()
